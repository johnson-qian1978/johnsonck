# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
import os

if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# 搜索文件
root = r"E:\工作目录"
apply_file = None
record_file = None

for dirpath, dirnames, filenames in os.walk(root):
    if '贵阳' in dirpath and '2025' in dirpath and '综合给付' in dirpath:
        if '最终数据备份' in dirpath:
            continue
        for filename in filenames:
            if filename.endswith('.xlsx') and not filename.startswith('~$'):
                fullpath = os.path.join(dirpath, filename)
                if '申请' in filename and '记录' not in filename:
                    apply_file = fullpath
                if '记录' in filename and '含作废' in filename:
                    record_file = fullpath

if not apply_file:
    print("ERROR: Could not find apply file")
    sys.exit(1)
if not record_file:
    print("ERROR: Could not find record file")
    sys.exit(1)

output_file = apply_file.replace('.xlsx', '_未找到标记.xlsx')

print(f"Apply file: {os.path.basename(apply_file)}")
print(f"Record file: {os.path.basename(record_file)}")
print(f"Output file: {os.path.basename(output_file)}")

# 读取 Excel 文件
print("\nReading apply file...")
df_apply = pd.read_excel(apply_file)
print(f"Columns: {list(df_apply.columns)}")
print(f"Rows: {len(df_apply)}")

print("\nReading record file...")
df_record = pd.read_excel(record_file)
print(f"Columns: {list(df_record.columns)}")
print(f"Rows: {len(df_record)}")

# 列名映射（申请文件 -> 记录文件）
# 申请文件：参加人员身份证号，住院开始日期，住院结束日期
# 记录文件：赔付人身份证号，住院日期，出院日期
apply_cols = {
    '身份证': '参加人员身份证号',
    '入院': '住院开始日期',
    '出院': '住院结束日期'
}
record_cols = {
    '身份证': '赔付人身份证号',
    '入院': '住院日期',
    '出院': '出院日期'
}

# 检查必要的列是否存在
for key, col in apply_cols.items():
    if col not in df_apply.columns:
        print(f"Error: Missing column '{col}' in apply file")
        print(f"Available columns: {list(df_apply.columns)}")
        sys.exit(1)

for key, col in record_cols.items():
    if col not in df_record.columns:
        print(f"Error: Missing column '{col}' in record file")
        print(f"Available columns: {list(df_record.columns)}")
        sys.exit(1)

# 创建记录集合用于快速查找
print("\nCreating record index...")
record_set = set()
for idx, row in df_record.iterrows():
    try:
        id_num = str(row.get(record_cols['身份证'], '')).strip()
        admit_date = str(row.get(record_cols['入院'], '')).strip()
        discharge_date = str(row.get(record_cols['出院'], '')).strip()
        key = (id_num, admit_date, discharge_date)
        record_set.add(key)
    except Exception as e:
        pass

print(f"Record set size: {len(record_set)}")

# 标记申请文件中找不到的行
print("\nComparing data...")
df_apply['found'] = False
for idx, row in df_apply.iterrows():
    id_num = str(row[apply_cols['身份证']]).strip()
    admit_date = str(row[apply_cols['入院']]).strip()
    discharge_date = str(row[apply_cols['出院']]).strip()
    key = (id_num, admit_date, discharge_date)
    if key in record_set:
        df_apply.loc[idx, 'found'] = True

# 统计结果
found_count = df_apply['found'].sum()
not_found_count = len(df_apply) - found_count
print(f"\nComparison result:")
print(f"  Found: {found_count} rows")
print(f"  Not found: {not_found_count} rows")

# 保存标记后的 Excel 文件（标黄未找到的行）
print(f"\nSaving results...")
wb = load_workbook(apply_file)
ws = wb.active
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 标记未找到的行（从第 2 行开始，因为第 1 行是标题）
for idx, row in df_apply.iterrows():
    if not row['found']:
        row_num = idx + 2
        for cell in ws[row_num]:
            cell.fill = yellow_fill

wb.save(output_file)
print(f"\nDone! Results saved to: {os.path.basename(output_file)}")
print(f"Highlighted {not_found_count} rows in yellow")
print(f"Full path: {output_file}")
